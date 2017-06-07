#
#   converter_excel.py
#
#   Description: This program is used to generate a formatted excel file from extracted data from .txt files in the current directory
#
#   Usage:
#   1. Put it in the folder with the .txt files
#   2. Open terminal
#   3. cd to the folder containing the .txt files and the .py file
#   4. Type command into terminal: python converter_excel.py
#   5. Cheers!
#

# This program will be used to enter
# specific .txt files into an excel format

import os
import glob
import numpy as np

import codecs

import errno

# replaced excel manipulation to use openpyxl
#import pandas as pd

import openpyxl
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
from openpyxl.styles import NamedStyle, Border, Side

#
#   Global variables
#   some a parameter, some b parameter, some c parameter in the text files. listLength is the length of the list of dictionaries
#
global_a = ''
global_b = ''
global_c = ''
global_listLength = 0

#   Name of the upper level/folder containing this .py file
containing_folder = ''

a_count = 0;
b_count = 0;
c_count = 0;

file_total_count = 0;
file_count = 0;

#
#   Custom fill template
#   highlight - green
#
highlight = PatternFill(start_color='7cfc00',
                      end_color='7cfc00',
                      fill_type='solid')

#
#   Custom fill template
#   highlight_error - red
#
highlight_error = PatternFill(start_color='f08080',
                        end_color='f08080',
                        fill_type='solid')

#
#   Custom border template
#   Stack Overflow
#   ws: worksheet name
#   cell_range: example input is 'A1:C5'
#   thickness: int from 1 to 3
#
def set_border(ws, cell_range, thickness):
    rows = ws[cell_range]
    if thickness == 1:
        side = Side(border_style='thin', color="FF000000")
    elif thickness == 2:
        side = Side(border_style='normal', color="FF000000")
    else:
        side = Side(border_style='thick', color="FF000000")
    
    rows = list(rows)  # we convert iterator to list for simplicity, but it's not memory efficient solution
    max_y = len(rows) - 1  # index of the last row
    for pos_y, cells in enumerate(rows):
        max_x = len(cells) - 1  # index of the last cell
        for pos_x, cell in enumerate(cells):
            border = Border(
                left=cell.border.left,
                right=cell.border.right,
                top=cell.border.top,
                bottom=cell.border.bottom
            )
            if pos_x == 0:
                border.left = side
            if pos_x == max_x:
                border.right = side
            if pos_y == 0:
                border.top = side
            if pos_y == max_y:
                border.bottom = side
                                                    
            # set new border only if it's one of the edge cells
            if pos_x == 0 or pos_x == max_x or pos_y == 0 or pos_y == max_y:
                cell.border = border

#####################
#                   #
#   Main Program    #
#                   #
#####################

# Get current working directory
cwd = os.getcwd()
print (cwd)

path,folder_name = os.path.split(os.getcwd())
containing_folder = folder_name

# To check if current directory is correct / lists files
files = [f for f in os.listdir('.') if os.path.isfile(f)]
for f in files:
    print(f)

# Testing code
#print ("")
#print ("--------------------")
# opening text file
#file = open("sample.txt","r")
#print (file.read())

# path test
print ()
path = cwd + "/*.txt"
print (path)

print ("\n-----------------------------")
print ("Finding and storing variables")
print ("-----------------------------\n")


# Initialize a list of dictionaries
dataList = [{"some a parameter":"test","some b parameter":"test","some c parameter":"test","File Name":"test"}]

# Testing code
#print (dataList[0]["some a parameter"])
#print (dataList[0]["some b parameter"])
#print (dataList[0]["some c parameter"])

# Keeps track of total files before skips in file data reading occur
for root, dirs, files in os.walk(cwd):
    for file in files:
        if file.endswith('.txt'):
            file_total_count += 1
#print (file_total_count)

# Go through each file
files = glob.glob(path)
for name in files:
    try:
        # codecs is added to ignore non-standard characters. ASCII crap
        with codecs.open(name, "r",encoding='utf-8',errors='ignore') as searchfile:
            # Go through each line in file
            for line in searchfile:
                left,sep,right = line.partition('a:')
                if sep: # True iff 'some a parameter' in line
                    global_a = right
                    global_a = global_a.strip()
                    #print(global_a)
                    break
            for line in searchfile:
                left,sep,right = line.partition('b:')
                if sep:
                    global_b = right
                    global_b = global_b.strip()
                    #print(global_b)
                    break
            for line in searchfile:
                left,sep,right = line.partition('c:')
                if sep:
                    global_c = right
                    global_c = global_c.strip()
                    #print(global_c)
                    break
            
            global_fileName = name

            # to separate data segments
            #print ()

            # store in dictionary
            dataList.append({"some a parameter":global_a,"some b parameter":global_b,"some c parameter":global_c,"File Name":global_fileName})

    except IOError as exc:
        if exc.errno != errno.EISDIR:
            raise

# Record list length
global_listLength = len(dataList)

#
#   Terminal tests
#

# For testing purposes
#print ()
#print ("List of dictionary results:")
#print ()

# Print testing
# Note: Uncomment to enable
#print("------------------")
#for i in range(0,len(dataList)):
#    objectName = "Object" + str(i)
#    print (objectName)
#    print (dataList[i]["some a parameter"])
#    print (dataList[i]["some b parameter"])
#    print (dataList[i]["some c parameter"])
#    print("------------------")

#
#   openpyxl usage
#

wb = Workbook()

dest_filename = 'empty_book.xlsx'

# Worksheet1 is the active workbook
ws1 = wb.active
ws1.title = "Spoof Data"

# Fills "title" cells for Worksheet1
ws1['A1'] = "some a parameter"
ws1['B1'] = "some b parameter"
ws1['C1'] = "some c parameter"
ws1['D1'] = "File Name"
ws1['E1'] = "Notes"

#
#   Reading from list of dictionaries into openpyxl methods
#

# Read values into column 1
for i in range(2, len(dataList)):
    ws1.cell(row=i, column=1).value = int(dataList[i]["some a parameter"])
    if int(dataList[i]["some a parameter"]) == 1:
        ws1.cell(row=i, column=1).fill = highlight
        a_count += 1

# Read values into column 2
for i in range(2, len(dataList)):
    ws1.cell(row=i, column=2).value = float(dataList[i]["some b parameter"])
    if float(dataList[i]["some b parameter"]) > 0.75:
        ws1.cell(row=i, column=2).fill = highlight
        b_count += 1
# Read values into column 3
for i in range(2, len(dataList)):
    if dataList[i]["some c parameter"] == "nan":
        ws1.cell(row=i, column=3).value = dataList[i]["some c parameter"]
    else:
        ws1.cell(row=i, column=3).value = float(dataList[i]["some c parameter"])
        if float(dataList[i]["some c parameter"]) > 0.25:
            ws1.cell(row=i, column=3).fill = highlight
            c_count += 1
    # Right justify cell for "nan" case
    ws1.cell(row=i, column=3).alignment = Alignment(horizontal='right')
# Read values into column 4
for i in range(2, len(dataList)):
    left,sep,right = dataList[i]["File Name"].partition(containing_folder)
    folder_directory = sep + right
    ws1.cell(row=i, column=4).value = sep + right
    
    # Workaround for counting "successful" files
    file_count += 1

#
#   Printing out values to terminal
#

# Print out values in column 1 to terminal
#for i in range(1, len(dataList)):
#    print(i, ws1.cell(row=i, column=1).value)
#print()
# Print out values in column 2 to terminal
#for i in range(1, len(dataList)):
#    print(i, ws1.cell(row=i, column=2).value)
#print()
# Print out values in column 3 to terminal
#for i in range(1, len(dataList)):
#    print(i, ws1.cell(row=i, column=3).value)
#print()

# Adjustments to column widths
ws1.column_dimensions["A"].width = 25.0
ws1.column_dimensions["B"].width = 25.0
ws1.column_dimensions["C"].width = 25.0
ws1.column_dimensions["D"].width = 75.0
ws1.column_dimensions["E"].width = 25.0
# Does not work
#dims = {}
#for row in range(1,2):#ws1.rows:
#    for cell in row:
#        if cell.value:
#            dims[cell.column] = max((dims.get(cell.column, 0), len(cell.value)))
#for col, value in dims.items():
#    ws1.column_dimensions[col].width = value

# Format cell borders
# Note: Just comment/remove below code to remove borders
set_border(ws1, "A1:C1", 3)
set_border(ws1, "B1:B1", 3)

custom_range = "A1:C" + str(len(dataList)-1)
set_border(ws1, custom_range, 3)

custom_range = "B1:B" + str(len(dataList)-1)
set_border(ws1, custom_range, 3)

set_border(ws1, "D1:D1", 3)
custom_range = "D1:D" + str(len(dataList)-1)
set_border(ws1, custom_range, 3)

set_border(ws1, "E1:E1", 3)
custom_range = "E1:E" + str(len(dataList)-1)
set_border(ws1, custom_range, 3)

# Borders for totals section
custom_range = "A" + str(len(dataList)) + ":E" + str(len(dataList))
set_border(ws1, custom_range, 3)
set_border(ws1, "B" + str(len(dataList)) + ":B" + str(len(dataList)), 3)
set_border(ws1, "D" + str(len(dataList)) + ":D" + str(len(dataList)), 3)

custom_range = "A" + str(len(dataList)+1) + ":E" + str(len(dataList)+1)
set_border(ws1, custom_range, 3)
set_border(ws1, "B" + str(len(dataList)+1) + ":B" + str(len(dataList)+1), 3)
set_border(ws1, "D" + str(len(dataList)+1) + ":D" + str(len(dataList)+1), 3)

# Count section
# Note: Total over thresholds. Conditions passed
ws1['A'+str(len(dataList)+1)] = a_count
ws1['B'+str(len(dataList)+1)] = b_count
ws1['C'+str(len(dataList)+1)] = c_count
ws1['A'+str(len(dataList)+1)].fill = highlight
ws1['B'+str(len(dataList)+1)].fill = highlight
ws1['C'+str(len(dataList)+1)].fill = highlight

# File count section
ws1['D'+str(len(dataList)+1)] = str(file_count) + "/" + str(file_total_count)
ws1['D'+str(len(dataList)+1)].alignment = Alignment(horizontal='right')
if file_count != file_total_count:
    ws1['D'+str(len(dataList)+1)].fill = highlight_error
else:
    ws1['D'+str(len(dataList)+1)].fill = highlight

#

# Save program changes
wb.save(filename = dest_filename)

print ("\n-----------------------------")
print ("Process completed")
print ("-----------------------------\n")
print ("Files successfully scanned:       " + str(file_count))
print ("Total files scanned in directory: " + str(file_total_count))

file_transfer = "";
if file_count == file_total_count:
    file_scan_status = "Process succeeded!";
else:
    file_scan_status = "Process failed!";
print ("Program status:                   " + file_scan_status)
print ()



















